#!/usr/bin/env python

import sys
import os
import glob
import csv
from collections import defaultdict
import openpyxl
import openpyxl.styles


def column_name(c):
    return chr((ord('A') + c-1))

def string_width(val):
    if isinstance(val, basestring):
        return len(val)
    else:
        return string_width(str(val))

def numberize(val):
    if isinstance(val, (int, long, float)):
        return val
    if isinstance(val, basestring):
        if '.' in val:
            try:
                return float(val)
            except ValueError:
                return val
        else:
            try:
                return int(val)
            except ValueError:
                try:
                    return long(val)
                except ValueError:
                    return val
    else:
        return numberize(str(va0))


def header_type(key):
    if key.startswith('%'): return '%'
    if key.startswith('$'): return '$'
    return ''

def main(output_filename):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    column_widths = defaultdict(int)
    for csvfile in glob.glob(os.path.join('.', '*.csv')):
        with open(csvfile, 'rb') as f:
            title = os.path.basename(csvfile).replace('.csv', '')
            worksheet.title = title
            reader = csv.reader(f, delimiter='\t')
            headers = None
            for r, row in enumerate(reader, start=1):
                if headers is None:
                    if not row[0].startswith('#'):
                        headers = [ header_type(key) for key in row ]
                for c, val in enumerate(row, start=1):
                    nv = numberize(val)
                    cell = worksheet.cell(row=r, column=c)
                    cell.value = nv
                    column_widths[c] = max(column_widths[c], string_width(nv))
        for col,column_width in column_widths.items():
            worksheet.column_dimensions[column_name(col)].width = column_width + 1
        worksheet = workbook.create_sheet()
        cols = set()
    workbook.save(output_filename)

if __name__ == "__main__":
  output_filename = sys.argv[1]
  main(output_filename)


